#!/usr/bin/env python3
"""Reusable experiment tracking & evaluation pipeline.

Creates/updates an Excel spreadsheet tracking every model modification.
For each run, it:
  1. Loads the model checkpoint
  2. Runs inference on story QA v4 plaintext shards
  3. Captures interpretable example outputs (story + Q&A completion)
  4. Records key metrics (val_loss, perplexity, params, etc.)
  5. Saves everything to the experiment tracker spreadsheet

Usage:
    # Record baseline
    python experiment_tracker.py \
        --run_name "baseline" \
        --modification "None (baseline SimpleTransformer)" \
        --intuition "Establish baseline performance" \
        --checkpoint checkpoints/baseline_full_8gpu.pt \
        --gradio_url "https://4de5f4801ccf07eacb.gradio.live" \
        --wandb_url "https://wandb.ai/jinyoungkim927/weightless/runs/dzpejlpz"

    # Record a future modification
    python experiment_tracker.py \
        --run_name "gqa_topk_v1" \
        --modification "Added GQA (4 kv heads) + TopK FFN (k=512)" \
        --intuition "Reduce KV cache and FFN inference cost" \
        --checkpoint checkpoints/gqa_topk_v1.pt
"""

import argparse
import os
import re
import json
import random
import datetime
from pathlib import Path

import torch
import torch.nn.functional as F
import numpy as np


# ============================================================================
# LLM-based semantic answer matching
# ============================================================================

def llm_judge_matches(qa_triples: list[dict], api_key: str = None) -> list[bool]:
    """Use Claude to judge whether model answers are semantically correct.

    Be very liberal — mark correct if meaning is conveyed, even with different
    phrasing, extra info, pronouns vs names, etc.
    """
    if not qa_triples:
        return []

    key = api_key or os.environ.get("ANTHROPIC_API_KEY", "")
    if not key:
        return _heuristic_matches(qa_triples)

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=key)

        items = []
        for i, t in enumerate(qa_triples):
            items.append(
                f"{i+1}. Question: {t['question']}\n"
                f"   Gold answer: {t['gold']}\n"
                f"   Model answer: {t['model']}"
            )

        prompt = (
            "You are judging whether a model's answer to a reading comprehension "
            "question is semantically correct compared to the gold answer.\n\n"
            "Be VERY liberal — the model answer is correct if it conveys the same "
            "meaning as the gold answer, even if:\n"
            "- It uses different words or phrasing\n"
            "- It's shorter or longer\n"
            "- It includes extra (but not contradictory) information\n"
            "- It paraphrases or rephrases the answer\n"
            "- It uses pronouns instead of names (or vice versa)\n\n"
            "Only mark as incorrect if the model answer is factually wrong, "
            "contradicts the gold, is completely off-topic, or is gibberish.\n\n"
            "Here are the items to judge:\n\n"
            + "\n\n".join(items) +
            "\n\nRespond with ONLY a JSON array of booleans, one per item. "
            "Example: [true, false, true]\n"
            "No explanation needed."
        )

        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}],
        )

        text = response.content[0].text.strip()
        if "```" in text:
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        results = json.loads(text)

        if isinstance(results, list) and len(results) == len(qa_triples):
            return [bool(r) for r in results]
        else:
            print(f"  ⚠ LLM returned unexpected format, falling back to heuristic")
            return _heuristic_matches(qa_triples)

    except Exception as e:
        print(f"  ⚠ LLM judge failed ({e}), falling back to heuristic")
        return _heuristic_matches(qa_triples)


def _heuristic_matches(qa_triples: list[dict]) -> list[bool]:
    """Simple heuristic fallback: substring + keyword overlap."""
    results = []
    for t in qa_triples:
        gold = t["gold"].lower().strip().rstrip(".")
        model = t["model"].lower().strip().rstrip(".")
        if gold in model or model in gold or gold == model:
            results.append(True)
            continue
        stop = {"a", "an", "the", "is", "was", "were", "are", "he", "she",
                "it", "they", "his", "her", "its", "to", "of", "in", "and",
                "that", "this", "yes", "no", "did", "do", "does"}
        gold_words = {w for w in gold.split() if w not in stop and len(w) > 1}
        model_words = set(model.split())
        if gold_words and len(gold_words & model_words) / len(gold_words) >= 0.6:
            results.append(True)
        else:
            results.append(False)
    return results

# Excel writing
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from transformers import GPT2Tokenizer

# Local modules
from model import create_model


# ============================================================================
# Story QA Data Loading
# ============================================================================

STORY_QA_DIR = os.path.join(os.path.dirname(__file__), "story_qa_v4_plaintext_shards")


def parse_story_qa_file(filepath: str) -> list[dict]:
    """Parse a story QA plaintext shard file into structured documents."""
    with open(filepath) as f:
        text = f.read()

    docs = []
    raw_docs = text.split("===== DOC START")
    for raw in raw_docs:
        if not raw.strip():
            continue

        header_match = re.search(r'split=(\w+)\s+idx=(\d+)\s+id=(\w+)', raw)
        if not header_match:
            continue
        split, idx, doc_id = header_match.groups()

        content = raw.split("=====\n", 1)[-1].split("===== DOC END")[0].strip()

        # Split story from QA pairs
        parts = content.split("<question>")
        story_text = parts[0].strip()
        # Remove inline definitions from story for cleaner display
        story_clean = re.sub(r'<definition>.*?</definition>', '', story_text).strip()
        story_clean = re.sub(r'\s+', ' ', story_clean)

        qa_pairs = []
        for p in parts[1:]:
            q_match = re.search(r'(.*?)</question>', p, re.DOTALL)
            a_match = re.search(r'<answer>(.*?)</answer>', p, re.DOTALL)
            if q_match and a_match:
                qa_pairs.append({
                    "question": q_match.group(1).strip(),
                    "answer": a_match.group(1).strip(),
                })

        definitions = re.findall(r'<definition>(.*?)</definition>', content)

        docs.append({
            "id": doc_id,
            "idx": int(idx),
            "split": split,
            "story": story_clean,
            "story_raw": story_text,
            "qa_pairs": qa_pairs,
            "definitions": definitions,
            "full_text": content,
        })

    return docs


def load_story_qa_examples(split="test", n_examples=10, seed=42) -> list[dict]:
    """Load a diverse random sample of story QA examples."""
    split_dir = os.path.join(STORY_QA_DIR, split)
    if not os.path.isdir(split_dir):
        raise FileNotFoundError(f"Story QA {split} dir not found: {split_dir}")

    all_docs = []
    for fname in sorted(os.listdir(split_dir)):
        if fname.endswith(".txt"):
            all_docs.extend(parse_story_qa_file(os.path.join(split_dir, fname)))

    random.seed(seed)
    # Pick a diverse sample: short stories, medium stories, long stories
    if len(all_docs) <= n_examples:
        return all_docs

    # Sort by story length and pick from different buckets
    all_docs_sorted = sorted(all_docs, key=lambda d: len(d["story"]))
    bucket_size = len(all_docs_sorted) // n_examples
    samples = []
    for i in range(n_examples):
        bucket = all_docs_sorted[i * bucket_size : (i + 1) * bucket_size]
        samples.append(random.choice(bucket))

    return samples


# ============================================================================
# Model Inference on Story QA
# ============================================================================

@torch.no_grad()
def run_story_qa_inference(model, tokenizer, device, examples: list[dict],
                           max_gen_tokens: int = 60, temperature: float = 0.7) -> list[dict]:
    """Run inference on story QA examples and capture interpretable outputs.

    For each example, we:
      1. Feed the story as context
      2. Try completing a question prompt
      3. Measure perplexity on the full doc
      4. Capture top token predictions
    """
    model.eval()
    results = []

    for ex in examples:
        story = ex["story"]
        qa = ex["qa_pairs"]

        if not qa:
            continue

        # Pick the first Q/A pair for completion test
        test_q = qa[0]
        prompt = f"{story}\n\nQuestion: {test_q['question']}\nAnswer:"

        # Encode
        input_ids = tokenizer.encode(prompt, return_tensors="pt").to(device)
        prompt_len = input_ids.shape[1]

        # Truncate if too long for context
        if prompt_len > 480:
            input_ids = input_ids[:, :480]
            prompt_len = 480

        # --- Measure perplexity on the story ---
        story_ids = tokenizer.encode(story, return_tensors="pt").to(device)
        if story_ids.shape[1] > 1:
            with torch.amp.autocast('cuda', dtype=torch.bfloat16):
                logits_story = model(story_ids)
            shift_logits = logits_story[:, :-1, :].contiguous()
            shift_labels = story_ids[:, 1:].contiguous()
            loss = F.cross_entropy(shift_logits.view(-1, shift_logits.size(-1)),
                                   shift_labels.view(-1)).item()
            ppl = float(np.exp(loss))
        else:
            loss = float('nan')
            ppl = float('nan')

        # --- Generate answer completion ---
        generated_ids = input_ids.clone()
        gen_tokens = []
        for _ in range(max_gen_tokens):
            with torch.amp.autocast('cuda', dtype=torch.bfloat16):
                logits = model(generated_ids)
            next_logits = logits[:, -1, :].float()

            # Get top predictions before sampling
            probs = F.softmax(next_logits, dim=-1)
            top5_probs, top5_ids = torch.topk(probs, 5, dim=-1)
            top5_tokens = [tokenizer.decode([tid]) for tid in top5_ids[0].tolist()]

            # Temperature sampling
            next_logits = next_logits / max(temperature, 1e-8)
            sampled = torch.multinomial(F.softmax(next_logits, dim=-1), num_samples=1)
            tok_id = sampled[0, 0].item()

            gen_tokens.append({
                "token": tokenizer.decode([tok_id]),
                "token_id": tok_id,
                "probability": probs[0, tok_id].item(),
                "top5": list(zip(top5_tokens, top5_probs[0].tolist())),
            })

            generated_ids = torch.cat([generated_ids, sampled], dim=1)

            # Stop on newline or EOS
            decoded_so_far = tokenizer.decode([tok_id])
            if tok_id == tokenizer.eos_token_id or "\n" in decoded_so_far:
                break

        generated_answer = tokenizer.decode(generated_ids[0, prompt_len:].tolist()).strip()

        results.append({
            "doc_id": ex["id"],
            "story": story[:500],  # truncate for spreadsheet
            "story_word_count": len(story.split()),
            "num_questions": len(qa),
            "test_question": test_q["question"],
            "gold_answer": test_q["answer"],
            "model_answer": generated_answer.split("\n")[0].strip(),
            "story_loss": loss,
            "story_perplexity": ppl,
            "gen_tokens": gen_tokens,
            "avg_confidence": np.mean([t["probability"] for t in gen_tokens]) if gen_tokens else 0,
            "has_definitions": len(ex["definitions"]) > 0,
        })

    return results


# ============================================================================
# 100-Sample Accuracy Benchmark
# ============================================================================

@torch.no_grad()
def run_100_sample_benchmark(model, tokenizer, device, seed=0,
                              max_gen_tokens=60, temperature=0.7,
                              n_samples=100) -> dict:
    """Run n_samples Story QA inferences and return accuracy dict."""
    model.eval()
    pool = []
    split_dir = os.path.join(STORY_QA_DIR, "test")
    if os.path.isdir(split_dir):
        for fname in sorted(os.listdir(split_dir)):
            if fname.endswith(".txt"):
                pool.extend(parse_story_qa_file(os.path.join(split_dir, fname)))
    pool = [d for d in pool if d["qa_pairs"]]
    if len(pool) < n_samples:
        print(f"  ⚠ Only {len(pool)} docs available, need {n_samples}")
        n_samples = len(pool)

    rng = random.Random(int(seed))
    samples = rng.sample(pool, n_samples)

    qa_triples = []
    for i, doc in enumerate(samples):
        if (i + 1) % 10 == 0 or i == 0:
            print(f"    Benchmark inference: {i+1}/{n_samples}")
        qa = doc["qa_pairs"][0]
        prompt = f'{doc["story"]}\n\nQuestion: {qa["question"]}\nAnswer:'
        input_ids = tokenizer.encode(prompt, return_tensors="pt").to(device)
        if input_ids.shape[1] > 480:
            input_ids = input_ids[:, :480]
        prompt_len = input_ids.shape[1]

        gen_ids = input_ids.clone()
        for _ in range(max_gen_tokens):
            with torch.amp.autocast('cuda', dtype=torch.bfloat16):
                logits = model(gen_ids)
            next_logits = logits[:, -1, :].float() / max(temperature, 1e-8)
            sampled = torch.multinomial(
                F.softmax(next_logits, dim=-1), num_samples=1)
            tok_id = sampled[0, 0].item()
            gen_ids = torch.cat([gen_ids, sampled], dim=1)
            if tok_id == tokenizer.eos_token_id or "\n" in tokenizer.decode([tok_id]):
                break

        model_answer = tokenizer.decode(
            gen_ids[0, prompt_len:].tolist()).strip().split("\n")[0]
        qa_triples.append({
            "question": qa["question"],
            "gold": qa["answer"],
            "model": model_answer,
        })

    # Batch LLM judge
    CHUNK = 25
    all_matches = []
    for start in range(0, len(qa_triples), CHUNK):
        all_matches.extend(llm_judge_matches(qa_triples[start:start + CHUNK]))

    n_correct = sum(all_matches)
    return {
        "n_correct": n_correct,
        "n_total": n_samples,
        "accuracy": n_correct / n_samples if n_samples > 0 else 0,
        "matches": all_matches,
    }


# ============================================================================
# Metrics Extraction
# ============================================================================

def extract_training_metrics(log_path: str = None) -> dict:
    """Extract final training metrics from train.log."""
    if log_path is None:
        log_path = os.path.join(os.path.dirname(__file__), "train.log")

    metrics = {}
    if not os.path.exists(log_path):
        return metrics

    with open(log_path) as f:
        text = f.read()

    # Final val loss
    m = re.search(r"Final val_loss:\s+([\d.]+)", text)
    if m:
        metrics["final_val_loss"] = float(m.group(1))

    # Training params
    m = re.search(r"TRAINING COMPLETE", text)
    metrics["training_complete"] = bool(m)

    # Goal achieved
    m = re.search(r"GOAL ACHIEVED", text)
    metrics["goal_achieved"] = bool(m)

    # Non-zero params
    m = re.search(r"([\d,]+)\s+non-zero params", text)
    if m:
        metrics["nonzero_params"] = int(m.group(1).replace(",", ""))

    # Total tokens trained — try multiple sources:

    # 1. Look in wandb summary JSON (most reliable)
    log_dir = os.path.dirname(os.path.abspath(log_path))
    wandb_dir = os.path.join(log_dir, "wandb")
    if os.path.isdir(wandb_dir):
        import glob
        summary_files = glob.glob(os.path.join(wandb_dir, "run-*/files/wandb-summary.json"))
        for sf in sorted(summary_files, reverse=True):  # newest first
            try:
                with open(sf) as f:
                    summary = json.load(f)
                if "train/total_tokens" in summary:
                    metrics["total_tokens_trained"] = int(summary["train/total_tokens"])
                    break
            except Exception:
                pass

    # 2. Fallback: look for inline pattern in log text
    if "total_tokens_trained" not in metrics:
        token_matches = re.findall(r"total_tokens['\"]?:\s*([\d]+)", text)
        if token_matches:
            metrics["total_tokens_trained"] = int(token_matches[-1])

    # 3. Last resort: calculate from config
    if "total_tokens_trained" not in metrics:
        bs_match = re.search(r"batch_size['\"]?:\s*(\d+)", text)
        ws_match = re.search(r"world_size\s+(\d+)", text)
        steps_match = re.search(r"num_steps['\"]?:\s*(\d+)", text)
        if bs_match and steps_match:
            bs = int(bs_match.group(1))
            ws = int(ws_match.group(1)) if ws_match else 1
            steps = int(steps_match.group(1))
            seq_len = 1023  # 1024 - 1 from collate_fn
            metrics["total_tokens_trained"] = bs * seq_len * ws * steps

    return metrics


def get_model_metrics(model) -> dict:
    """Extract metrics directly from the model."""
    total_params = model.count_parameters(count_zeros=True)
    nonzero_params = model.count_parameters(count_zeros=False)

    return {
        "total_params": total_params,
        "nonzero_params": nonzero_params,
        "sparsity": 1 - nonzero_params / total_params if total_params > 0 else 0,
        "d_model": model.d_model,
        "n_layers": model.n_layers,
        "n_heads": model.n_heads,
        "n_kv_heads": model.n_kv_heads,
        "d_ff": model.d_ff,
    }


# ============================================================================
# Excel Spreadsheet Creation & Update
# ============================================================================

SPREADSHEET_PATH = os.path.join(os.path.dirname(__file__), "experiment_tracker.xlsx")

# Column widths
COL_WIDTHS = {
    "A": 22, "B": 45, "C": 45, "D": 20, "E": 14, "F": 14,
    "G": 18, "H": 14, "I": 18, "J": 18,
    "K": 25, "L": 25, "M": 55, "N": 55,
}

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
METRIC_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
EXAMPLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def create_or_load_workbook(path: str = None):
    """Create or load the experiment tracker workbook."""
    path = path or SPREADSHEET_PATH
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Experiments"
        _setup_experiments_sheet(ws)
    return wb


def _setup_experiments_sheet(ws):
    """Set up the main experiments tracking sheet with headers."""
    headers = [
        "Run Name",           # A
        "Modification",       # B
        "Intuition",          # C
        "Training Run",       # D
        "Val Loss",           # E
        "Perplexity",         # F
        "Non-zero Params",    # G
        "Sparsity %",         # H
        "# Tokens Trained",   # I
        "Story QA Acc %",     # J
        "Checkpoint Path",    # K
        "Gradio URL",         # L
        "WandB URL",          # M
        "Timestamp",          # N
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Set column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze top row
    ws.freeze_panes = "A2"


def _setup_examples_sheet(ws, run_name: str):
    """Set up headers for a per-run examples sheet."""
    headers = [
        "Doc ID",             # A
        "Story (truncated)",  # B
        "Words",              # C
        "# Questions",        # D
        "Test Question",      # E
        "Gold Answer",        # F
        "Model Answer",       # G
        "Match?",             # H
        "Story Loss",         # I
        "Story PPL",          # J
        "Avg Confidence",     # K
        "Top-5 for 1st Token",# L
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    widths = {"A": 18, "B": 60, "C": 8, "D": 10, "E": 40, "F": 40,
              "G": 40, "H": 10, "I": 12, "J": 12, "K": 14, "L": 60}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"


def add_experiment_row(wb, run_name, modification, intuition, metrics,
                       checkpoint_path, gradio_url, wandb_url,
                       training_run_info="", story_qa_accuracy=None):
    """Add a row to the Experiments sheet."""
    ws = wb["Experiments"]
    row = ws.max_row + 1

    val_loss = metrics.get("final_val_loss", metrics.get("val_loss", ""))
    ppl = np.exp(val_loss) if isinstance(val_loss, (int, float)) else ""
    nonzero = metrics.get("nonzero_params", "")
    if isinstance(nonzero, (int, float)):
        nonzero_display = f"{nonzero:,}"
    else:
        nonzero_display = str(nonzero)
    sparsity = metrics.get("sparsity", "")
    if isinstance(sparsity, (int, float)):
        sparsity_display = f"{sparsity:.2%}"
    else:
        sparsity_display = str(sparsity)

    total_tokens = metrics.get("total_tokens_trained", "")
    if isinstance(total_tokens, (int, float)) and total_tokens > 0:
        tokens_display = f"{total_tokens / 1e9:.2f}B"
    else:
        tokens_display = str(total_tokens)

    if story_qa_accuracy is not None:
        acc_display = f"{story_qa_accuracy:.1%}"
    else:
        acc_display = ""

    values = [
        run_name,                                          # A
        modification,                                       # B
        intuition,                                          # C
        training_run_info,                                  # D
        val_loss,                                           # E
        ppl,                                                # F
        nonzero_display,                                    # G
        sparsity_display,                                   # H
        tokens_display,                                     # I
        acc_display,                                        # J
        checkpoint_path,                                    # K
        gradio_url,                                         # L
        wandb_url,                                          # M
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M"), # N
    ]

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER
        # Colour-code metric cells
        if col_idx in (5, 6, 7, 8, 9, 10):
            cell.fill = METRIC_FILL

    return row


def add_examples_sheet(wb, run_name, inference_results: list[dict]):
    """Create a per-run sheet with spot-checked inference examples."""
    sheet_name = f"Examples_{run_name}"[:31]  # Excel 31-char limit

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)
    _setup_examples_sheet(ws, run_name)

    for i, r in enumerate(inference_results):
        row = i + 2

        # Format top-5 for first generated token
        top5_str = ""
        if r.get("gen_tokens"):
            t = r["gen_tokens"][0]
            top5_str = " | ".join(f"{tok}({prob:.0%})" for tok, prob in t["top5"])

        match_str = "✅" if r.get("is_match") else "❌"

        values = [
            r["doc_id"],
            r["story"][:300],
            r["story_word_count"],
            r["num_questions"],
            r["test_question"],
            r["gold_answer"],
            r["model_answer"],
            match_str,
            f"{r['story_loss']:.4f}" if not np.isnan(r["story_loss"]) else "N/A",
            f"{r['story_perplexity']:.2f}" if not np.isnan(r["story_perplexity"]) else "N/A",
            f"{r['avg_confidence']:.2%}",
            top5_str,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER
            if col_idx in (9, 10, 11):
                cell.fill = METRIC_FILL
            elif col_idx in (5, 6, 7, 8):
                cell.fill = EXAMPLE_FILL


def add_data_overview_sheet(wb):
    """Add a sheet summarizing the story QA dataset."""
    sheet_name = "Dataset_Overview"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    # Title
    ws.merge_cells("A1:E1")
    cell = ws.cell(row=1, column=1, value="Story QA v4 Plaintext Shards – Dataset Overview")
    cell.font = Font(bold=True, size=14, color="2F5496")

    # Stats
    stats = [
        ("Dataset", "year1-story-qa"),
        ("Version", "v4 (train) / v4_eval (test)"),
        ("Train Documents", "100,000"),
        ("Test Documents", "2,048"),
        ("Train Shards", "10 files (~16 MB each)"),
        ("Test Shards", "1 file (~3.1 MB)"),
        ("Total Train Questions", "~1,038,314"),
        ("Total Test Questions", "~21,432"),
        ("Avg Questions/Doc", "~10.4"),
        ("Avg Story Length", "~80 words (~445 chars)"),
        ("Docs with Definitions", "~90.6%"),
        ("", ""),
        ("Question Type Distribution (test set)", ""),
        ("  What-questions", "62.8%"),
        ("  Where-questions", "9.1%"),
        ("  Who-questions", "6.8%"),
        ("  Did-questions", "6.4%"),
        ("  How-questions", "6.0%"),
        ("  Was-questions", "5.2%"),
        ("  Other", "3.7%"),
        ("", ""),
        ("Format", "Plaintext with <question>, <answer>, <definition> tags"),
        ("Document Delimiter", "===== DOC START / DOC END ====="),
    ]

    row = 3
    for label, value in stats:
        ws.cell(row=row, column=1, value=label).font = Font(bold=bool(label and not label.startswith(" ")))
        ws.cell(row=row, column=2, value=value)
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 55


# ============================================================================
# Main: Full Pipeline
# ============================================================================

def run_full_pipeline(
    run_name: str,
    modification: str,
    intuition: str,
    checkpoint_path: str = None,
    model_variant: str = "baseline",
    d_model: int = 768,
    n_layers: int = 8,
    n_heads: int = 8,
    d_ff: int = 2048,
    gradio_url: str = "",
    wandb_url: str = "",
    n_examples: int = 10,
    log_path: str = None,
    extra_model_kwargs: dict = None,
):
    """Full pipeline: load model → run inference → update spreadsheet.

    This function is designed to be called after each training run / modification.
    """
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    tokenizer = GPT2Tokenizer.from_pretrained("gpt2")

    print(f"\n{'='*60}")
    print(f"  Experiment Tracker: {run_name}")
    print(f"{'='*60}")

    # 1. Load model
    print(f"\n[1/5] Loading model ({model_variant}) ...")
    model_kwargs = dict(
        variant=model_variant,
        d_model=d_model,
        n_layers=n_layers,
        n_heads=n_heads,
        d_ff=d_ff,
    )
    if extra_model_kwargs:
        model_kwargs.update(extra_model_kwargs)
    model = create_model(**model_kwargs)

    if checkpoint_path and os.path.exists(checkpoint_path):
        sd = torch.load(checkpoint_path, map_location=device, weights_only=True)
        model.load_state_dict(sd)
        print(f"  ✓ Loaded checkpoint: {checkpoint_path}")
    else:
        print(f"  ⚠ No checkpoint found at {checkpoint_path} — using random-init.")

    model.to(device)
    model.eval()

    # 2. Get model metrics
    print("\n[2/5] Extracting model metrics ...")
    model_metrics = get_model_metrics(model)
    train_metrics = extract_training_metrics(log_path)
    all_metrics = {**model_metrics, **train_metrics}
    print(f"  val_loss: {all_metrics.get('final_val_loss', 'N/A')}")
    print(f"  non-zero params: {all_metrics.get('nonzero_params', 'N/A'):,}")

    # 3. Load story QA examples
    print(f"\n[3/5] Loading {n_examples} story QA examples ...")
    examples = load_story_qa_examples(split="test", n_examples=n_examples)
    print(f"  Loaded {len(examples)} examples from test set")
    for ex in examples[:3]:
        print(f"    {ex['id']}: {ex['story'][:80]}...")

    # 4. Run inference
    print(f"\n[4/5] Running inference on story QA examples ...")
    inference_results = run_story_qa_inference(model, tokenizer, device, examples)
    print(f"  Completed {len(inference_results)} inferences")

    # 4b. LLM-based semantic matching
    qa_triples = [
        {"question": r["test_question"],
         "gold": r["gold_answer"],
         "model": r["model_answer"]}
        for r in inference_results
    ]
    judge_method = "LLM" if os.environ.get("ANTHROPIC_API_KEY") else "heuristic"
    print(f"  Judging matches ({judge_method}) ...")
    match_results = llm_judge_matches(qa_triples)
    for r, m in zip(inference_results, match_results):
        r["is_match"] = m

    n_match = sum(match_results)
    print(f"  Match rate: {n_match}/{len(match_results)} "
          f"({n_match/len(match_results):.0%})" if match_results else "")

    # Print a few example results
    for r in inference_results[:3]:
        match_icon = "✅" if r.get("is_match") else "❌"
        print(f"\n  --- {r['doc_id']} {match_icon} ---")
        print(f"  Q: {r['test_question']}")
        print(f"  Gold: {r['gold_answer']}")
        print(f"  Model: {r['model_answer']}")
        print(f"  Story PPL: {r['story_perplexity']:.2f}, Avg Conf: {r['avg_confidence']:.2%}")

    # 4c. 100-sample accuracy benchmark
    print(f"\n[4c] Running 100-sample accuracy benchmark ...")
    bench_results = run_100_sample_benchmark(model, tokenizer, device)
    story_qa_accuracy = bench_results["accuracy"]
    print(f"  100-sample accuracy: {bench_results['n_correct']}/100 = {story_qa_accuracy:.1%}")

    # 5. Update spreadsheet
    print(f"\n[5/5] Updating spreadsheet: {SPREADSHEET_PATH}")
    wb = create_or_load_workbook()

    # Add dataset overview (only first time)
    if "Dataset_Overview" not in wb.sheetnames:
        add_data_overview_sheet(wb)
        print("  ✓ Added dataset overview sheet")

    # Training run info
    training_info = f"{'COMPLETE' if all_metrics.get('training_complete') else 'IN PROGRESS'}"
    if all_metrics.get("goal_achieved"):
        training_info += " ✓ GOAL"

    add_experiment_row(
        wb, run_name, modification, intuition, all_metrics,
        checkpoint_path or "", gradio_url, wandb_url, training_info,
        story_qa_accuracy=story_qa_accuracy,
    )
    print(f"  ✓ Added experiment row for '{run_name}'")

    add_examples_sheet(wb, run_name, inference_results)
    print(f"  ✓ Added examples sheet 'Examples_{run_name}'")

    wb.save(SPREADSHEET_PATH)
    print(f"\n  ✅ Saved: {SPREADSHEET_PATH}")
    print(f"{'='*60}\n")

    return {
        "metrics": all_metrics,
        "inference_results": inference_results,
        "spreadsheet_path": SPREADSHEET_PATH,
    }


# ============================================================================
# CLI
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Experiment tracker & evaluation pipeline")
    parser.add_argument("--run_name", required=True, help="Short name for this run")
    parser.add_argument("--modification", required=True,
                        help="What was changed from baseline")
    parser.add_argument("--intuition", required=True,
                        help="Why this modification was made")
    parser.add_argument("--checkpoint", default=None,
                        help="Path to model checkpoint")
    parser.add_argument("--model", default="baseline",
                        choices=["baseline", "baseline_plus", "copy_gate"])
    parser.add_argument("--d_model", type=int, default=768)
    parser.add_argument("--n_layers", type=int, default=8)
    parser.add_argument("--n_heads", type=int, default=8)
    parser.add_argument("--d_ff", type=int, default=2048)
    parser.add_argument("--gradio_url", default="", help="Gradio app URL")
    parser.add_argument("--wandb_url", default="", help="WandB run URL")
    parser.add_argument("--n_examples", type=int, default=10,
                        help="Number of story QA examples to evaluate")
    parser.add_argument("--log_path", default=None,
                        help="Path to train.log for metric extraction")
    parser.add_argument("--ffn_type", default=None,
                        help="FFN type: swiglu or relu2 (passed to create_model)")
    parser.add_argument("--qk_norm", action="store_true", default=False,
                        help="Enable QK normalization (passed to create_model)")
    parser.add_argument("--softcap", type=float, default=0.0,
                        help="Logit softcapping value (passed to create_model)")
    parser.add_argument("--resid_scalars", action="store_true", default=False,
                        help="Enable per-layer residual scalars (passed to create_model)")

    args = parser.parse_args()

    extra_model_kwargs = {}
    if args.ffn_type:
        extra_model_kwargs["ffn_type"] = args.ffn_type
    if args.qk_norm:
        extra_model_kwargs["qk_norm"] = True
    if args.softcap > 0:
        extra_model_kwargs["softcap"] = args.softcap
    if args.resid_scalars:
        extra_model_kwargs["use_resid_scalars"] = True

    run_full_pipeline(
        run_name=args.run_name,
        modification=args.modification,
        intuition=args.intuition,
        checkpoint_path=args.checkpoint,
        model_variant=args.model,
        d_model=args.d_model,
        n_layers=args.n_layers,
        n_heads=args.n_heads,
        d_ff=args.d_ff,
        gradio_url=args.gradio_url,
        wandb_url=args.wandb_url,
        n_examples=args.n_examples,
        log_path=args.log_path,
        extra_model_kwargs=extra_model_kwargs or None,
    )


if __name__ == "__main__":
    main()
